import openpyxl
import json

wb = openpyxl.load_workbook(r'e:\trae\2LIMS\报价单.xlsx')
print(f"Sheet names: {wb.sheetnames}\n")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*60}")
    print(f"Sheet: {sheet_name}")
    print(f"{'='*60}")
    
    # Get dimensions
    print(f"Max row: {ws.max_row}, Max col: {ws.max_column}\n")
    
    # Print first 30 rows
    for i, row in enumerate(ws.iter_rows(max_row=30, values_only=True), 1):
        # Filter out completely empty rows
        if any(cell is not None for cell in row):
            print(f"Row {i:2d}: {row}")
